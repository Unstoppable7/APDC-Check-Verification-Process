import pandas as pd
import os #Manejo de directorios y archivos
import sys
from openpyxl import load_workbook,Workbook
import time
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

import threading

pd.set_option('mode.chained_assignment', None)
OUTPUT_DIRECTORY = "OUTPUT"
DATA_BEFORE_PROCESSING_DIRECTORY = "DATA BEFORE PROCESSING"
PROCESSED_DATA_DIRECTORY = "PROCESSED DATA"
TRANSACTIONS_TO_REVIEW_DIRECTORY = "TRANSACTIONS TO REVIEW"
FORMATTED_BANK_REPORT_NAME = "BANK REPORT.xlsx"
FORMATTED_QB_REPORT_NAME = "QUICKBOOKS REPORT.xlsx"
PENDING_TRANSACTIONS_BEFORE_PROCESSING_NAME = "PENDING TRANSACTIONS BEFORE PROCESSING.xlsx"
DUPLICATE_PENDING_TRANSACTIONS_NAME = "DUPLICATE PENDING TRANSACTIONS.xlsx"
VOID_TRANSACTIONS_NAME = "VOID TRANSACTIONS.xlsx"
CONFIRMED_TRANSACTIONS_NAME = "CONFIRMED TRANSACTIONS.xlsx"
PENDING_TRANSACTIONS_PROCESSED_NAME = "PENDING TRANSACTIONS PROCESSED.xlsx"
RESULT_TRANSACTIONS_PROCESSED_NAME = "RESULT - TRANSACTIONS PROCESSED.xlsx"
PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME = " QB PENDING"
CONFIRMED_TRANSACTIONS_PAGE_MAIN_FILE_NAME = " QB CONFIRMED"
BANK_PENDING_TRANSACTIONS_NAME = "BANK PENDING TRANSACTIONS.xlsx"

# Crear una clase personalizada para redirigir la salida a la caja de texto
class TextRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, msg):
        self.text_widget.insert(tk.END, msg)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

def abrir_archivo_bank():
    archivo = filedialog.askopenfilename(title="Seleccionar archivo Bank Report", filetypes=[("Archivos XLSX", "*.xlsx"), ("Archivos CSV", "*.csv")])
    if archivo:
        entrada_bank.delete(0, tk.END)
        entrada_bank.insert(0, archivo)

def abrir_archivos_quickbooks():
    archivos = filedialog.askopenfilenames(title="Seleccionar archivos QuickBooks Reports", filetypes=[("Archivos XLSX", "*.xlsx")])
    if archivos:
        entrada_quickbooks.delete(0, tk.END)
        entrada_quickbooks.insert(0, "<<>> ".join(archivos))

def show_info_Quickbooks_report():
     messagebox.showinfo("Info", "Select the report files extracted from Quickbooks with check transactions\n\nEach file must represent an account, which will be searched in the TD Bank transactions report\n\nThis version uses the report extracted from Quickbooks Reports - Memorized Reports - Zuleika")

    #  text = "\n\nIMPORTANT: The exported Quickbooks file must contain only transactions that have not been verified, that is, when exporting the report, filtering must be done by check number from the last check that has not yet been processed"

def show_info_bank_report():
     messagebox.showinfo("Info", "Select the report file extracted from TD Bank \nYou can use a multi-account or single-account file\n\nIn this file the accounts will be searched according to Quickbooks Transactions Reports")
def show_info_main_file():
     messagebox.showinfo("Info", "History file of all pending transactions and confirmed transactions\n\nThis file will be updated according to the confirmed transactions and the pending transactions resulting from the process.")

def abrir_archivo_central():
    archivo = filedialog.askopenfilename(title="Seleccionar archivo central", filetypes=[("Archivos XLSX", "*.xlsx")])
    if archivo:
        entrada_archivo_central.delete(0, tk.END)
        entrada_archivo_central.insert(0, archivo)

def update_progress_bar(i, total_tareas):
     # Calcular el progreso basado en el número de tareas completadas
    progreso_a_sumar = (i) / total_tareas * 100

    # Actualizar la barra de progreso
    barra_progreso["value"] = barra_progreso["value"] + progreso_a_sumar
    ventana.update_idletasks()  # Actualizar la ventana para mostrar el progreso

def procesar():
    #Limpiamos la caja de texto de informacion
    clean(clean_inputs=False)
    
    #Variables que manejan la barra de progreso
    reading_files = 3
    format_reports = 2
    new_pendings = 2
    matching_pendings = 2
    insert_data = 2    

    ##################################### Manejo de entrada de archivos
    archivo_bank = entrada_bank.get()
    archivos_quickbooks = entrada_quickbooks.get()
    archivo_central = entrada_archivo_central.get()
    #Tratamos si los archivos estan vacios
    if archivo_bank == "": 
        messagebox.showerror("Error","No file has been selected as bank report")
        return
    if archivos_quickbooks == "":
        messagebox.showerror("Error", "No file selected as Quickbooks report")
        return
    if archivo_central == "":
        messagebox.showerror("Error", "No file selected as main file")
        return
    
    archivo_central_name = archivo_central.rsplit('/', 1)[1]
    #Mostramos barra de progreso
    toggle_progress_bar(True)

    qb_reports_dict = {}
    if(archivo_bank.rsplit('.', 1)[1] == 'csv'):

        barra_progreso_label.config(text=f"Converting {archivo_bank.rsplit('/', 1)[1]} to xlsx file")
        ventana.update_idletasks()

        try:
            df = pd.read_csv(archivo_bank)
            
        except PermissionError:
            messagebox.showerror("Error", f"Permission denied when accessing file '{archivo_bank.rsplit('/', 1)[1]}'. \n\nIf you have the file open please close it and try again")
            clean(False)
            return
        except FileNotFoundError:
            messagebox.showerror("Error", f"The file '{archivo_bank.rsplit('/', 1)[1]}' was not found. \n\nVerify that you selected the correct file and try again")
            clean(False)
            return
        except Exception as e:
            messagebox.showerror("Error", f"{e} \n\nError converting CSV file to Excel.\n\nPlease check {archivo_bank.rsplit('/', 1)[1]} file")
            clean(False)
            return
        
        df.to_excel(archivo_bank.rsplit('.', 1)[0] + ".xlsx", index=False)    
        try:
            os.remove(archivo_bank)
        except:
            pass
        archivo_bank = archivo_bank.rsplit('.', 1)[0] + ".xlsx"

        barra_progreso["value"] = 100
        time.sleep(0.5)    
    
    barra_progreso_label.config(text="Reading files")
    barra_progreso["value"] = 0
    ventana.update_idletasks()

    qb_reports_account_number_file_name_dict = {}
    archivos_quickbooks_list = archivos_quickbooks.split('<<>> ')
    for file in archivos_quickbooks_list:
        
        # Verificacion de formato quickbook report
        try:
            tmp_dt = pd.read_excel(file, header=None, sheet_name="Sheet1")           

        except:
            messagebox.showerror("Error", "The Quickbooks report '" + file.rsplit('/', 1)[1] +"' is not in the correct format. \n\nIt is necessary that the sheet where the report is located has the name 'Sheet1'. \n\nPlease check the Quickbooks report and try again")
            clean(clean_inputs=False)
            return

        try:
            if not (str(tmp_dt.iat[0, 4]) == "Type" and str(tmp_dt.iat[0, 6]) == "Date" and str(tmp_dt.iat[0, 8]) == "Num" and str(tmp_dt.iat[0, 10]) == "Name" and str(tmp_dt.iat[0, 12]) == "Debit" and str(tmp_dt.iat[0, 14]) == "Credit" and str(tmp_dt.iat[0, 16])) == "Balance":
                messagebox.showerror("Error", "The Quickbooks report '" + file.rsplit('/', 1)[1] +"' is not in the correct format. \n\nPlease check the Quickbooks report and try again")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(clean_inputs=False)
                return
        except:
            messagebox.showerror("Error", "It is not possible to access the information of the Quickbooks report '" + file.rsplit('/', 1)[1] +"' \n\nPlease check the Quickbooks report and try again")
            clean(False)
            return

        #Diccionario que guarda el data frame, cada llave son los ultimos cuatro digitos del numero de cuenta
        qb_reports_dict[str(tmp_dt.iat[1, 1])[:4]] = tmp_dt
        #Diccionario para poder mostrar el mensaje con el numero de cuenta y el nombre del archivo
        qb_reports_account_number_file_name_dict[str(tmp_dt.iat[1, 1])[:4]] = file.rsplit('/', 1)[1]

        #Actualizamos la barra de progreso
        update_progress_bar(1/len(archivos_quickbooks_list),reading_files)

    bank_full_report = pd.read_excel(archivo_bank, header=None)

    #Verificacion formato TDBank report
    if not (str(bank_full_report.iat[0, 0]) == "Date" and str(bank_full_report.iat[0, 1]) == "Bank RTN" and str(bank_full_report.iat[0, 2]) == "Account Number" and str(bank_full_report.iat[0, 3]) == "Transaction Type" and str(bank_full_report.iat[0, 4]) == "Description" and str(bank_full_report.iat[0, 5]) == "Debit" and str(bank_full_report.iat[0, 6]) == "Credit" and str(bank_full_report.iat[0, 7]) == "Check Number" and str(bank_full_report.iat[0, 8])) == "Account Running Balance":
        messagebox.showerror("Error", "The bank report '" + archivo_bank.rsplit('/', 1)[1] +"' is not in the correct format. \n\nPlease check the bank report and try again")
        #Ocultamos barra de progreso y limpiamos caja de texto
        clean(False)
        return

    bank_full_report[2] = bank_full_report[2].astype(str)

    bank_reports = []

    #Extraemos todas las llaves del diccionario donde almacenamos los reportes de quickbooks
    accounts_to_look_for = qb_reports_dict.keys()
    for account_to_look_for in accounts_to_look_for:
        # Filtrar las filas basadas en el valor específico
        if bank_full_report[bank_full_report[2].str[-4:] == account_to_look_for].empty:

            messagebox.showinfo("Warning", "The account number " + account_to_look_for + " extracted from the '" + qb_reports_account_number_file_name_dict[account_to_look_for] +"' file is not found in the bank report '" + archivo_bank.rsplit('/', 1)[1] + "'")
        else:
            bank_reports.append(bank_full_report[bank_full_report[2].str[-4:] == account_to_look_for])
        
        #Actualizamos la barra de progreso
        update_progress_bar(1/len(accounts_to_look_for),reading_files)
    
    account_numbers_separate_by_comma = ", ".join(str(elemento) for elemento in accounts_to_look_for)
    if len(bank_reports) <= 0:
        messagebox.showwarning("Warning", f"No matches found for accounts number {account_numbers_separate_by_comma} in the bank report " + archivo_bank.rsplit('/', 1)[1] + "'")
        #Ocultamos barra de progreso y limpiamos caja de texto
        clean(False)
        return

    try:
        #Accedemos al archivo principal
        main_file = load_workbook(filename = archivo_central)
    except:
        messagebox.showerror("Error",f"Could not access file {archivo_central_name} \nMake sure you have selected the correct file")

        clean(clean_inputs=False)
        return
    #Actualizamos barra de progreso
    barra_progreso["value"] = 100
    ventana.update_idletasks()

    accounts_processed = []

    retry = True
    ##################################### Bucle de cuentas a procesar
    for i in range(len(bank_reports)):
        
        time.sleep(1)

        #Actualizamos label barra de progreso
        barra_progreso_label.config(text="Formatting reports")
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 0
        ventana.update_idletasks()

        if bank_reports[i].empty:
            messagebox.showwarning("Error", f"There is a problem extracting the data from the bank report\nPlease check {archivo_bank.rsplit('/', 1)[1]}")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        bank_report = bank_reports[i].copy()

        #Accedemos manualmente a la celda donde se encuentra el numero de cuenta
        #Convertimos a string y tomamos los ultimos 4 digitos
        account_number_bank_report = str(bank_report.iat[1, 2])[-4:]
        
        print(f"Processing account {account_number_bank_report}\n")
        ventana.update_idletasks()

        #Extraemos el reporte de quickbooks acorde a la cuenta extraida del banco
        qb_report = qb_reports_dict[account_number_bank_report].copy()

        os.makedirs(os.path.join(OUTPUT_DIRECTORY,account_number_bank_report,DATA_BEFORE_PROCESSING_DIRECTORY), exist_ok=True)
        os.makedirs(os.path.join(OUTPUT_DIRECTORY,account_number_bank_report,PROCESSED_DATA_DIRECTORY), exist_ok=True)
        ##################################### Formato reportes

        # Eliminar columnas
        columnas_a_eliminar_qb_report = [0,1,2,3,4,5,7,9,11,12,13,15,16]  # Índices de las columnas a eliminar
        columnas_extras_a_eliminar = 4 #A partir del este numero se eliminaran
        qb_report = qb_report.drop(qb_report.columns[columnas_a_eliminar_qb_report], axis=1)

        if qb_report.shape[1] > 4:
            qb_report = qb_report.drop(qb_report.columns[columnas_extras_a_eliminar:], axis=1)
            
        # Eliminar columnas
        columnas_a_eliminar_bank_report = [1,2,3,4,6,8]  # Índices de las columnas a eliminar
        columnas_extras_a_eliminar = 3 #A partir del este numero se eliminaran

        bank_report = bank_report.drop(bank_report.columns[columnas_a_eliminar_bank_report], axis=1)

        if bank_report.shape[1] > 3:
            bank_report = bank_report.drop(bank_report.columns[columnas_extras_a_eliminar:], axis=1)

        # Especificar el índice de la columna en la que deseas verificar los valores nulos
        indice_columna_verificacion_qb_report = 1

        # Eliminar las filas con valores nulos en la columna especificada
        qb_report = qb_report.dropna(subset=[qb_report.columns[indice_columna_verificacion_qb_report]])

        # Especificar el índice de la columna en la que deseas verificar los valores nulos
        indice_columna_verificacion_bank_report = 2  

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)

        # Eliminar las filas con valores nulos en la columna especificada
        bank_report = bank_report.dropna(subset=[bank_report.columns[indice_columna_verificacion_bank_report]])

        #Verificamos que el reporte del banco ya formateado no este vacio
        if bank_report.empty:
            rsp = messagebox.askyesno("Warning", f"Account number {account_number_bank_report} does not have check-type transactions in '" + archivo_bank.rsplit('/', 1)[1] + "' file\n\nDo you want to process this account anyway?")
            
            if not rsp:

                print(f"Account {account_number_bank_report} was not processed\n")
                continue

        # Eliminar filas
        filas_a_eliminar = [0]  # Índices de las filas a eliminar
        qb_report = qb_report.drop(filas_a_eliminar)

        #Reseteamos el nombre de las columnas del qb_report 
        qb_report.columns = range(qb_report.shape[1])
        bank_report.columns = range(bank_report.shape[1])

        try:
            # Convertir las columnas a tipo de datos de fecha
            qb_report[0] = pd.to_datetime(qb_report[0])
            # Formatear las fechas en el DataFrame
            qb_report[0] = qb_report[0].dt.strftime('%m/%d/%Y')

            # Convertir las columnas a tipo de datos de fecha
            bank_report[0] = pd.to_datetime(bank_report[0])

            # Formatear las fechas en el DataFrame
            bank_report[0] = bank_report[0].dt.strftime('%m/%d/%Y')
        except Exception as e:
            messagebox.showerror("Error", f"{e}\n\nCould not format dates on bank reports or Quicbooks reports")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)

        ####FIN#### FORMATO LIMPIO QB REPORT, BANK REPORT ########
        # Guardar los cambios
        try_again = True
        while try_again:
            try:
                qb_report.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{DATA_BEFORE_PROCESSING_DIRECTORY}/{FORMATTED_QB_REPORT_NAME}', index=False, header=None)
            
                try_again = False
            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {FORMATTED_QB_REPORT_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + FORMATTED_QB_REPORT_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return
        
        try_again = True
        while try_again:
            try:
                bank_report.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{DATA_BEFORE_PROCESSING_DIRECTORY}/{FORMATTED_BANK_REPORT_NAME}', index=False, header=None)

                try_again = False
            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {FORMATTED_BANK_REPORT_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + FORMATTED_BANK_REPORT_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return
        
        #Agregamos columna que identifica de donde proviene la transaccion
        qb_report.insert(4, 'FROM', "QUICKBOOKS REPORT")
        
        #Actualizamos la barra de progreso
        update_progress_bar(1,format_reports)
        
        time.sleep(1)

        #Actualizamos label barra de progreso
        barra_progreso_label.config(text="Unifying pending transactions")
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 0
        ventana.update_idletasks()
        ##################################### Extraccion pending archivo central

        try:
            #Leemos con pandas el archivo central
            old_transaction_pending_data_frame = pd.read_excel(archivo_central, header=None, sheet_name=account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME)
        except ValueError:
            messagebox.showerror("Error", "Page '" + account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' not found in '" + archivo_central_name + "'\n\nPlease verify that '" + archivo_central_name + "' contains  '" + account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        except Exception as e:
            messagebox.showerror("Error", f"{e}\n\nPlease check {archivo_central_name} and {account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME} sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return
        #Actualizamos la barra de progreso
        update_progress_bar(1/2,new_pendings)

        #Verificacion de formato de pendings en archivo central
        if len(old_transaction_pending_data_frame.columns) != 4:
            messagebox.showerror("Error", "The page '" + account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' of file '" + archivo_central_name + "' does not have the expected format. \n\nPlease check '" + archivo_central_name +"' file and try again")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        #Agregamos columna que identifica de donde proviene la transaccion
        old_transaction_pending_data_frame.insert(4, 'FROM', archivo_central_name.upper())

        ##################################### Unificacion pendings

        #Concatenamos para tener todos los pending en un solo dataframe
        current_transactions_pending_unprocessed = pd.concat([qb_report, old_transaction_pending_data_frame], ignore_index=True)

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)

        # Aplicar la transformación a las columnas a entero para que realice correctamente el drop_duplicates
        current_transactions_pending_unprocessed[1] = current_transactions_pending_unprocessed[1].astype(int)
        current_transactions_pending_unprocessed[2] = current_transactions_pending_unprocessed[2].astype(str)
        current_transactions_pending_unprocessed[3] = current_transactions_pending_unprocessed[3].astype(float)

        #Eliminamos duplicados con el mismo nro de transaccion y el mismo monto (dejamos una sola transaccion)
        current_transactions_pending_unprocessed.drop_duplicates(inplace=True,subset=[1,2,3],keep='first')

        # Extraemos un df con True en las filas duplicadas y False en las que no
        duplicate_transactions = current_transactions_pending_unprocessed.duplicated(subset=[1],keep=False)

        # Extraer filas duplicadas
        duplicate_transactions = current_transactions_pending_unprocessed[duplicate_transactions]

        # Ordenamos por numero de transaccion
        duplicate_transactions = duplicate_transactions.sort_values(by=[1,3],axis=0)

        # Eliminamos todos los duplicados
        current_transactions_pending_unprocessed.drop_duplicates(inplace=True,subset=[1],keep=False)

        # Extraemos transacciones sin monto (void)
        transaction_amount_column = 3
        void_transactions = current_transactions_pending_unprocessed.loc[current_transactions_pending_unprocessed[current_transactions_pending_unprocessed.columns[transaction_amount_column]].isnull()]

        #Eliminamos las transacciones en las que el monto sea vacio (void)
        current_transactions_pending_unprocessed = current_transactions_pending_unprocessed.dropna(subset=[current_transactions_pending_unprocessed.columns[transaction_amount_column]])

        #Le damos a la columna 1 el tipo de objeto datetime
        current_transactions_pending_unprocessed[0] = pd.to_datetime(current_transactions_pending_unprocessed[0])

        #Le damos a la columna 1 el tipo de objeto datetime
        duplicate_transactions[0] = pd.to_datetime(duplicate_transactions[0])

        #Le damos a la columna 1 el tipo de objeto datetime
        void_transactions[0] = pd.to_datetime(void_transactions[0])

        # Ordenamos por fecha 
        current_transactions_pending_unprocessed = current_transactions_pending_unprocessed.sort_values(by=[0],axis=0)
        
        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)

        try:
            # Formatear las fechas en el DataFrame
            current_transactions_pending_unprocessed[0] = current_transactions_pending_unprocessed[0].dt.strftime('%m/%d/%Y')

            # Formatear las fechas en el DataFrame
            duplicate_transactions[0] = duplicate_transactions[0].dt.strftime('%m/%d/%Y')

            # Formatear las fechas en el DataFrame
            void_transactions[0] = void_transactions[0].dt.strftime('%m/%d/%Y')
        except Exception as e:
            messagebox.showerror("Error", str(e) + "\n\nCould not format the dates of '" + PENDING_TRANSACTIONS_BEFORE_PROCESSING_NAME + "' or '" + DUPLICATE_PENDING_TRANSACTIONS_NAME + "' or '" + VOID_TRANSACTIONS_NAME + "'")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        try_again = True
        while try_again:
            try:
                #Guardamos las transacciones unicas y pendientes
                current_transactions_pending_unprocessed.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{DATA_BEFORE_PROCESSING_DIRECTORY}/{PENDING_TRANSACTIONS_BEFORE_PROCESSING_NAME}', index=False, header=False)  
                
                try_again = False
            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {PENDING_TRANSACTIONS_BEFORE_PROCESSING_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + PENDING_TRANSACTIONS_BEFORE_PROCESSING_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return        
        
        try_again = True
        while try_again:
            try:
                #Guardamos los duplicados en un archivo aparte
                if(not duplicate_transactions.empty):
                    #Creamos el directorio
                    os.makedirs(os.path.join(OUTPUT_DIRECTORY,account_number_bank_report,TRANSACTIONS_TO_REVIEW_DIRECTORY), exist_ok=True)

                    duplicate_transactions.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{TRANSACTIONS_TO_REVIEW_DIRECTORY}/{DUPLICATE_PENDING_TRANSACTIONS_NAME}', header=None, index=False)  
                    print(f"\tDuplicate pending transactions file has been created\n")

                try_again = False

            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {DUPLICATE_PENDING_TRANSACTIONS_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return
            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" +  DUPLICATE_PENDING_TRANSACTIONS_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return

        try_again = True
        while try_again:
            try:
                #Guardamos las transacciones void en un archivo aparte
                if(not void_transactions.empty):
                    #Creamos el directorio
                    os.makedirs(os.path.join(OUTPUT_DIRECTORY,account_number_bank_report,TRANSACTIONS_TO_REVIEW_DIRECTORY), exist_ok=True)

                    void_transactions.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{TRANSACTIONS_TO_REVIEW_DIRECTORY}/{VOID_TRANSACTIONS_NAME}', header=None, index=False)  
                    print(f"\tVoid transactions file has been created\n")
                
                try_again = False

            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {VOID_TRANSACTIONS_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return
            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + VOID_TRANSACTIONS_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return
            
        ventana.update_idletasks()

        #Eliminamos la columna que agregamos para identificar desde donde proviene la transaccion
        current_transactions_pending_unprocessed = current_transactions_pending_unprocessed.drop(current_transactions_pending_unprocessed.columns[4], axis=1)

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)
        
        time.sleep(1)

        #Actualizamos label barra de progreso
        barra_progreso_label.config(text="Matching transactions")
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 0
        ventana.update_idletasks()
        ##################################### Matching pendings y bank report

        # Realizar la búsqueda y transferir los datos
        transactions_processed = pd.merge(current_transactions_pending_unprocessed, bank_report, left_on=1, right_on=2, how='outer', sort=True, indicator=True)

        #Movemos la columna donde se encuentra el numero de transaccion en el area del reporte bancario
        #Para que quede en la posicion esperada
        position = 6  # Índice de la posición deseada
        columns = transactions_processed.columns.tolist()
        column_to_move = columns[7]  # Índice de la columna que deseas mover (en este ejemplo, la tercera columna)
        columns.remove(column_to_move)
        columns.insert(position, column_to_move)
        transactions_processed = transactions_processed[columns]

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,matching_pendings)

        # Filtrar las filas basadas en el valor específico
        bank_transactions_pending = transactions_processed[transactions_processed['_merge'] == 'right_only']

        #Le damos formato a bank_transactions_pending
        columns_to_delete = [0,1,2,3,4,8]
        bank_transactions_pending = bank_transactions_pending.drop(bank_transactions_pending.columns[columns_to_delete], axis=1)

        try_again = True
        while try_again:
            try:
                #Guardamos las transacciones pendientes del banco en un archivo aparte
                if(not bank_transactions_pending.empty):
                    #Creamos el directorio
                    os.makedirs(os.path.join(OUTPUT_DIRECTORY,account_number_bank_report,TRANSACTIONS_TO_REVIEW_DIRECTORY), exist_ok=True)

                    bank_transactions_pending.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{TRANSACTIONS_TO_REVIEW_DIRECTORY}/{BANK_PENDING_TRANSACTIONS_NAME}', index=False,header=None)  
                    print(f"\tBank pending transactions file has been created\n")

                try_again = False
                
            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {BANK_PENDING_TRANSACTIONS_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + BANK_PENDING_TRANSACTIONS_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return
        ventana.update_idletasks()

        # Crear un nuevo DataFrame con las filas no filtradas
        transactions_processed = transactions_processed[transactions_processed['_merge'] != 'right_only']

        #Formato al archivo final
        #Eliminar primera columna y columna del indicator _merge
        transactions_processed = transactions_processed.drop([transactions_processed.columns[0],transactions_processed.columns[8]], axis=1)

        # Obtener filas no coincidentes
        transactions_pending_processed = transactions_processed[transactions_processed.isnull().any(axis=1)]

        #Le damos a la columna 1 el tipo de objeto datetime
        transactions_pending_processed['0_x'] = pd.to_datetime(transactions_pending_processed['0_x'])

        # Ordenamos por numero de transaccion 
        transactions_pending_processed['1_x'] = transactions_pending_processed['1_x'].astype(int)
        transactions_pending_processed = transactions_pending_processed.sort_values(by=['1_x'],axis=0)
        
        # Formatear las fechas en el DataFrame
        transactions_pending_processed['0_x'] = transactions_pending_processed['0_x'].dt.strftime('%m/%d/%Y')

        try_again = True
        while try_again:        
            try:
                transactions_pending_processed.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{PROCESSED_DATA_DIRECTORY}/{PENDING_TRANSACTIONS_PROCESSED_NAME}', index=False, header=None)  
                
                try_again = False

            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {PENDING_TRANSACTIONS_PROCESSED_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + PENDING_TRANSACTIONS_PROCESSED_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return

        #Agregamos la columnas
        transactions_pending_processed.insert(0, 'STATUS', "PENDING")
        transactions_pending_processed.insert(3, 'MATCH', "")
        transactions_pending_processed.insert(4, 'QB COMMENTS', "")
        transactions_pending_processed.insert(10, 'TEST #', "")
        transactions_pending_processed.insert(11, 'TEST AMOUNT', "")
        transactions_pending_processed.insert(12, 'BANK COMMENTS', "")
        transactions_pending_processed.insert(13, 'EXTRA', "")

        # Mover filas no coincidentes al final del archivo
        confirmed_transactions = transactions_processed.dropna()
        transactions_processed = transactions_processed.dropna()

        #Le damos formato fecha a la columna correspondiente
        confirmed_transactions['0_x'] = pd.to_datetime(confirmed_transactions['0_x'])

        # Ordenamos por fecha 
        confirmed_transactions = confirmed_transactions.sort_values(by=['0_x'],axis=0)

        # Formatear las fechas en el DataFrame
        confirmed_transactions['0_x'] = confirmed_transactions['0_x'].dt.strftime('%m/%d/%Y')

        confirmed_transactions.insert(2, 'MATCH', "")
        confirmed_transactions.insert(3, 'QB COMMENTS', "")

        try_again = True
        while try_again:
            try:    
                # Guardamos las filas coincidentes
                confirmed_transactions.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{PROCESSED_DATA_DIRECTORY}/{CONFIRMED_TRANSACTIONS_NAME}', index=False, header=None)  

                try_again = False
            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {CONFIRMED_TRANSACTIONS_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + CONFIRMED_TRANSACTIONS_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return

        #Agregamos la columna status
        transactions_processed.insert(0, 'STATUS', "READY")
        transactions_processed.insert(3, 'MATCH', "")
        transactions_processed.insert(4, 'QB COMMENTS', "")
        transactions_processed.insert(10, 'TEST #', "")
        transactions_processed.insert(11, 'TEST AMOUNT', "")
        transactions_processed.insert(12, 'BANK COMMENTS', "")
        transactions_processed.insert(13, 'EXTRA', "")

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,matching_pendings)

        transactions_processed = pd.concat([transactions_processed, transactions_pending_processed])

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,matching_pendings)

        transactions_processed.columns = ['STATUS', 'QB DATE', 'QB NUMBER', 'MATCH', 'QB COMMENTS', 'NAME', 'QB AMOUNT', 'BANK DATE', 'BANK NUMBER', 'BANK AMOUNT','TEST #', 'TEST AMOUNT', 'BANK COMMENTS', 'EXTRA']

        try_again = True
        while try_again:
            try:
                #Guardamos archivo con todos los datos finales
                transactions_processed.to_excel(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{PROCESSED_DATA_DIRECTORY}/{RESULT_TRANSACTIONS_PROCESSED_NAME}', index=False)  

                try_again = False

            except PermissionError:
                
                rsp = messagebox.askretrycancel("Permission error", f"Could not update file {RESULT_TRANSACTIONS_PROCESSED_NAME} \nIf you have this file open please close it \n\nDo you want to try again?")

                if not rsp:
                    clean(False)
                    return

            except Exception as e:
                messagebox.showerror("Error", str(e) + "\n\nFailed to export file '" + RESULT_TRANSACTIONS_PROCESSED_NAME + "'")
                #Ocultamos barra de progreso y limpiamos caja de texto
                clean(False)
                return
            
        #Actualizamos la barra de progreso
        update_progress_bar(1/2,format_reports)
        
        time.sleep(1)

        #Actualizamos label barra de progreso
        barra_progreso_label.config(text=f"Updating {archivo_central_name}")
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 0   
        ventana.update_idletasks()
        ##################################### Insercion de resultado en archivo central

        try:
            #Abrimos el archivo con las transacciones confirmadas procesadas
            confirmed_transactions = load_workbook(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{PROCESSED_DATA_DIRECTORY}/{CONFIRMED_TRANSACTIONS_NAME}')

            #Actualizamos la barra de progreso
            update_progress_bar(1/2,insert_data)

            #Abrimos el archivo con las transacciones pendientes procesadas
            transactions_pending_processed = load_workbook(f'{OUTPUT_DIRECTORY}/{account_number_bank_report}/{PROCESSED_DATA_DIRECTORY}/{PENDING_TRANSACTIONS_PROCESSED_NAME}')

        except Exception as e:
            messagebox.showerror("Error", str(e) + "\n\nThere is a problem with '" + CONFIRMED_TRANSACTIONS_NAME + "' or '" + PENDING_TRANSACTIONS_PROCESSED_NAME + "'")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,insert_data)

        try:
            #Accedemos a la pestaña con las transacciones confirmadas
            confirmed_transactions_page_main_file = main_file[account_number_bank_report + CONFIRMED_TRANSACTIONS_PAGE_MAIN_FILE_NAME]
        except KeyError:
            messagebox.showerror("Error", "Page '" + account_number_bank_report + CONFIRMED_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' not found in '" + archivo_central_name + "'\n\nPlease verify that '" + archivo_central_name + "' contains '" + account_number_bank_report + CONFIRMED_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return
        except Exception as e:
            messagebox.showerror("Error", f"{e}\n\nPlease check {archivo_central_name} and {account_number_bank_report + CONFIRMED_TRANSACTIONS_PAGE_MAIN_FILE_NAME} sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return
        try:
            #Accedemos a la pestaña con las transacciones pendientes
            pending_transactions_page_main_file = main_file[account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME]
        except KeyError:
            messagebox.showerror("Error", "Page '" + account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' not found in '" + archivo_central_name + "'\n\nPlease verify that '" + archivo_central_name + "' contains '" + account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME + "' sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return
        except Exception as e:
            messagebox.showerror("Error", f"{e}\n\nPlease check {archivo_central_name} and {account_number_bank_report + PENDING_TRANSACTIONS_PAGE_MAIN_FILE_NAME} sheet")
            #Ocultamos barra de progreso y limpiamos caja de texto
            clean(False)
            return
        #Extraemos los datos del archivo de transacciones confirmadas
        confirmed_transactions_rows = list(confirmed_transactions.active.iter_rows(values_only=True))

        process_approved = False
        while(not process_approved and retry):

            try:
                # Insertamos los datos en el archivo original
                for row in confirmed_transactions_rows:
                    confirmed_transactions_page_main_file.append(row)

                process_approved = True
            except PermissionError:
                
                retry = messagebox.askretrycancel("Error inserting confirmed transactions", f"Could not access file {archivo_central_name} \nIt is probably open \n\nDo you want to try again?")

            except Exception as e:
                                
                retry = messagebox.askretrycancel("Error", f"{e}\n\nPlease check {archivo_central_name}\n\nDo you want to try again?")

        #Extraemos los datos del archivo de transacciones pendientes
        pending_transactions_rows = list(transactions_pending_processed.active.iter_rows(values_only=True))
        
        #Actualizamos la barra de progreso
        update_progress_bar(1/2,insert_data)

        process_approved = False
        while(not process_approved and retry):

            try:
                #Borramos las transacciones pendientes anteriores
                pending_transactions_page_main_file.delete_rows(1, pending_transactions_page_main_file.max_row)


                process_approved = True

            except PermissionError:

                retry = messagebox.askretrycancel("Error deleting old pending transactions", f"Could not access file {archivo_central_name} \nIt is probably open \n\nDo you want to try again?")

            except Exception as e:
                
                retry = messagebox.askretrycancel("Error", f"{e}\n\nPlease check {archivo_central_name}\n\nDo you want to try again?")

        
        process_approved = False
        while(not process_approved and retry):

            try:
                for row in pending_transactions_rows:
                    pending_transactions_page_main_file.append(row)


                process_approved = True
                
            except PermissionError:

                retry = messagebox.askretrycancel("Error inserting updated pending transactions", f"Could not access file {archivo_central_name} \nIt is probably open \n\nDo you want to try again?")

            except Exception as e:
                
                retry = messagebox.askretrycancel("Error", f"{e}\n\nPlease check {archivo_central_name}\n\nDo you want to try again?")
    
        if not retry:
            messagebox.showwarning("Warning", f"Account {account_number_bank_report} will not be updated on {archivo_central_name}\n\n In the folder {OUTPUT_DIRECTORY}/{account_number_bank_report} you can find the processed results")
            continue

        #Actualizamos la barra de progreso
        update_progress_bar(1/2,insert_data)

        print(f"Account {account_number_bank_report} processing completed\n")

        accounts_processed.append(account_number_bank_report)

        ventana.update_idletasks()

    process_approved = False
    while(not process_approved and retry and (len(accounts_processed) > 0)):

        #Actualizamos label barra de progreso
        barra_progreso_label.config(text=f"Saving {archivo_central.rsplit('/',1)[1]}")
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 0   
        ventana.update_idletasks()

        time.sleep(1)
        #Reiniciamos barra de progreso
        barra_progreso["value"] = 50   
        ventana.update_idletasks()

        try:
            main_file.save(archivo_central)
            
            process_approved = True
        except PermissionError:

            retry = messagebox.askretrycancel("Error saving file", f"Could not access file {archivo_central_name} \nIt is probably open \n\nDo you want to try again?")

        except Exception as e:

            retry = messagebox.askretrycancel("Error", f"{e}\n\nPlease check {archivo_central_name}\n\nDo you want to try again?")
    
    if not retry:
        messagebox.showwarning("Warning", f"The file {archivo_central_name} was not saved\n\n In the folder {OUTPUT_DIRECTORY} you can find the processed results")
        #Ocultamos barra de progreso
        toggle_progress_bar(False)
        return


    #Reiniciamos barra de progreso
    barra_progreso["value"] = 100   
    ventana.update_idletasks()

    time.sleep(1)

    #Ocultamos barra de progreso
    toggle_progress_bar(False)
    
    if len(accounts_processed) > 0 :
        messagebox.showinfo("Success", "The following accounts have been processed\n\n" + "\n".join(str(elemento) for elemento in accounts_processed))
    else:
        messagebox.showwarning("Warning", "No account could be processed")

    print("Finished process")

def clean(clean_inputs = True):
    caja_texto.delete(1.0, tk.END)
    if clean_inputs:
        entrada_quickbooks.delete(0, tk.END)
        entrada_archivo_central.delete(0, tk.END)
        entrada_bank.delete(0, tk.END)

    #Ocultamos barra de progreso
    toggle_progress_bar(False)

def toggle_progress_bar(band):
    if band:
        #Barra de progreso
        barra_progreso_label.place(relx=0.5, rely=0.5, anchor="center")
        barra_progreso.place(relx=0.5, rely=0.55, anchor="center")
    else:   
        # Actualizar la barra de progreso
        barra_progreso["value"] = 0
        barra_progreso_label.config(text="")
        barra_progreso_label.place_forget()
        barra_progreso.place_forget()
    
    ventana.update_idletasks()

def ejecutar_hilo():
    # Crear un objeto Thread y pasarle la función procesar como objetivo
    hilo_procesar = threading.Thread(target=procesar)

    # Iniciar la ejecución del hilo
    hilo_procesar.start()

    
# Crear la ventana principal
ventana = tk.Tk()

# Configurar el tamaño de la ventana
ventana.geometry("1200x600")

# Configurar el mensaje de bienvenida
mensaje_bienvenida = tk.Label(ventana, text="APDC Check verification process", font=("Arial", 24))
# mensaje_bienvenida.grid(row=0, column=0, columnspan=2, pady=20)

# Configurar los subtitulos y las cajas de texto de Bank Report
subtitulo_bank = tk.Label(ventana, text="TD Bank Transactions Report", font=("Arial", 18))
# subtitulo_bank.grid(row=1, column=0, padx=20, sticky="w")

entrada_bank = tk.Entry(ventana, width=40)
# entrada_bank.grid(row=2, column=0, padx=20)

boton_bank = tk.Button(ventana, text="Select file", command=abrir_archivo_bank)
boton_info_bank = tk.Button(ventana, text="Info", command=show_info_bank_report)

# boton_bank.grid(row=3, column=0, padx=20, pady=10)

# Configurar los subtitulos y las cajas de texto de QuickBooks Report
subtitulo_quickbooks = tk.Label(ventana, text="QuickBooks Transactions Reports", font=("Arial", 18))
# subtitulo_quickbooks.grid(row=1, column=1, padx=20, sticky="w")

entrada_quickbooks = tk.Entry(ventana, width=40)
# entrada_quickbooks.grid(row=2, column=1, padx=20)

boton_quickbooks = tk.Button(ventana, text="Select files", command=abrir_archivos_quickbooks)
boton_info_quickbooks = tk.Button(ventana, text="Info", command=show_info_Quickbooks_report)
# boton_quickbooks.grid(row=3, column=1, padx=20, pady=10)

# Configurar el botón de procesamiento
boton_procesar = tk.Button(ventana, text="Start Process", font=("Arial", 22), command=ejecutar_hilo)
# boton_procesar.grid(row=4, column=0, columnspan=2, pady=20)

# Configurar la caja de texto
caja_texto = tk.Text(ventana, height=10, width=100)
# caja_texto.grid(row=5, column=0, columnspan=2, padx=20)

# Configurar el botón "Clean"
boton_clean = tk.Button(ventana, text="Clean", font=("Arial", 14), command=clean)
# boton_clean.grid(row=6, column=0, columnspan=2, pady=10)

# Crear una barra de progreso
barra_progreso = ttk.Progressbar(ventana, mode="determinate", length=300)
barra_progreso_label = tk.Label(ventana, text="", font=("Arial", 12))

subtitulo_archivo_central = tk.Label(ventana, text="Main File", font=("Arial", 18))
entrada_archivo_central = tk.Entry(ventana, width=40)
boton_archivo_central = tk.Button(ventana, text="Select file", command=abrir_archivo_central)
boton_info_main_file = tk.Button(ventana, text="Info", command=show_info_main_file)

#Place
mensaje_bienvenida.place(relx=0.5, rely=0.1, anchor="center")

# Configurar los subtitulos y las cajas de texto de archivo central
subtitulo_archivo_central.place(relx=0.5, rely=0.2, anchor="center")
entrada_archivo_central.place(relx=0.5, rely=0.25, anchor="center")
boton_archivo_central.place(relx=0.5, rely=0.3, anchor="center")
boton_info_main_file.place(relx=0.62, rely=0.25, anchor="center")

# Configurar los subtitulos y las cajas de texto de Bank Report
subtitulo_bank.place(relx=0.8, rely=0.2, anchor="center")
entrada_bank.place(relx=0.8, rely=0.25, anchor="center")
boton_bank.place(relx=0.8, rely=0.3, anchor="center")
boton_info_bank.place(relx=0.92, rely=0.25, anchor="center")

# Configurar los subtitulos y las cajas de texto de QuickBooks Report
subtitulo_quickbooks.place(relx=0.2, rely=0.2, anchor="center")
entrada_quickbooks.place(relx=0.2, rely=0.25, anchor="center")
boton_quickbooks.place(relx=0.2, rely=0.3, anchor="center")
boton_info_quickbooks.place(relx=0.32, rely=0.25, anchor="center")

# Configurar el botón de procesamiento
boton_procesar.place(relx=0.5, rely=0.4, anchor="center")

# #Barra de progreso
# barra_progreso_label.place(relx=0.5, rely=0.5, anchor="center")
# barra_progreso.place(relx=0.5, rely=0.55, anchor="center")

# Configurar la caja de texto
caja_texto.place(relx=0.5, rely=0.742, anchor="center")

# Configurar el botón "Clean"
boton_clean.place(relx=0.5, rely=0.95, anchor="center")

# Redirigir la salida estándar y la salida de error a la caja de texto
sys.stdout = TextRedirector(caja_texto)
sys.stderr = TextRedirector(caja_texto)

# # Establecer el icono de la ventana
# icono = 'APDC LOGO.ico'  # Reemplaza con la ruta completa del archivo de icono
# ventana.iconbitmap(icono)
ventana.title("APDC Check verification process")

# Ejecutar el bucle principal de la interfaz gráfica
ventana.mainloop()